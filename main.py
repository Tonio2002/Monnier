from fastapi import FastAPI, File, UploadFile
import pandas as pd
import re
import shutil
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import os  # Pour gérer les fichiers
import uvicorn

# Importer python-multipart pour la gestion des fichiers en multipart
import multipart

app = FastAPI()

@app.post("/format-excel/")
async def format_excel(file: UploadFile = File(...)):
    # 📌 Sauvegarde temporaire du fichier
    file_location = f"/tmp/{file.filename}"
    with open(file_location, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # 📌 Charger le fichier Excel avec Pandas
    xls = pd.ExcelFile(file_location)
    df = pd.read_excel(xls, sheet_name="Report1")

    # 📌 Étape 1 : Ajouter la colonne "Site Name"
    site_names = []
    current_site = None
    for index, row in df.iterrows():
        if isinstance(row.iloc[0], str) and " - " in row.iloc[0]:
            current_site = row.iloc[0]  
        site_names.append(current_site)
    df.insert(0, "Site Name", site_names)

    # 📌 Étape 2 : Supprimer la ligne 4 (index 3)
    df = df.drop(index=3).reset_index(drop=True)

    # 📌 Étape 3 : Ajouter une colonne vide en B
    df.insert(1, "New Column", "")

    # 📌 Étape 4 : Déplacer le contenu des parenthèses en colonne B
    df["New Column"] = df["Retail Sales"].apply(
        lambda x: re.search(r"\((.*?)\)", x).group(1) if isinstance(x, str) and "(" in x else ""
    )

    # 📌 Étape 5 : Supprimer les parenthèses et leur contenu en colonne originale
    df["Retail Sales"] = df["Retail Sales"].str.replace(r"\s*\(.*?\)", "", regex=True)

    # 📌 Étape 6 : Sauvegarde du fichier transformé
    output_file_path = "/tmp/RetailSales_Final.xlsx"
    df.to_excel(output_file_path, index=False)

    # 📌 Étape 7 : Appliquer la mise en forme (colorier les lignes contenant "Total")
    wb = load_workbook(output_file_path)
    ws = wb.active
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "Total" in cell.value:
                for cell_in_row in row:
                    cell_in_row.fill = fill  # Colorier toute la ligne en jaune

    # 📌 Sauvegarder avec la mise en forme
    wb.save(output_file_path)

    # 📌 Retourner l'URL de téléchargement
    return {"download_url": "http://localhost:8000/download/RetailSales_Final.xlsx"}
